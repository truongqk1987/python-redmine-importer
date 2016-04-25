from openpyxl import load_workbook
from redmine import Redmine


def get_worksheet(workbook_name, worksheet_name):
    wb = load_workbook(workbook_name)
    ws = wb[worksheet_name]
    return ws


def get_redmine(redmine_url, username, password):
    redmine = Redmine(redmine_url, username=username, password=password)
    return redmine

def get_redmine_project(redmine, project_id):
    project = redmine.project.get(project_id) # can be id or string
    return project;


def import_excel_to_redmine(ws, filtered_range, project_id, redmine):
    feature_id = None
    NGUYEN_VAN_VUI = 12 # Back-End developer
    TRUONG_QUANG_KHAI = 11 # Front-End developer
    for row in ws.iter_rows(filtered_range):
        feature_name = row[0].value
        task_name = row[1].value
        front_end_estimation = row[2].value
        back_end_estimation = row[3].value
        if (feature_name is not None):
            issue = create_redmine_issue(redmine, ws, project_id, feature_name, 2, 1)
            feature_id = issue.id
        elif (feature_name is None):
            if (task_name is None and front_end_estimation is None and back_end_estimation is None):
                print('Please remove blank')
            else:
                if (front_end_estimation is not None):
                    user_id = TRUONG_QUANG_KHAI
                    estimated_hours = front_end_estimation
                else:
                    user_id = NGUYEN_VAN_VUI
                    estimated_hours = back_end_estimation
                issue = create_redmine_issue(redmine, ws, project_id, task_name, 4, 1, feature_id, user_id, estimated_hours)
           
            
def create_redmine_issue(redmine, ws, project_id, issue_name, tracker_id, status_id, parent_issue_id=None, assigned_to_id=None,  estimated_hours=None):
    issue = redmine.issue.create(
        project_id = project_id,
        subject = issue_name,
        tracker_id = tracker_id,
        status_id = status_id,
        assigned_to_id = assigned_to_id,
        parent_issue_id = parent_issue_id,
        estimated_hours = estimated_hours
    )
    return issue
    
    
def let_it_go():
    username = '<your_user_name>'
    password = '<your_password>'
    redmine_url = '<your_redmine>'
    workbook_name = '<excel_file_name>'
    worksheet_name = '<worksheet_name>' # For example: My Sheet
    project_id = '<project_id>' # For example: my-project
    filtered_range = '' # For example: A1:D389
    redmine = get_redmine(redmine_url, username, password)
    ws = get_worksheet(workbook_name, worksheet_name)
    project = get_redmine_project(redmine, project_id)
    import_excel_to_redmine(ws, filtered_range, project_id, redmine)
    #delete_all_issues(redmine, project_id)

def delete_all_issues(redmine, project_id):
    for issue in redmine.issue.filter(project_id=project_id):
        redmine.issue.delete(issue.id)
    
let_it_go()